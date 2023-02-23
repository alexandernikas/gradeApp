#!/usr/bin/env python
# coding: utf-8

# In[2]:


#import openpyxl
#import shutil
import os
try:
    import openpyxl
except ImportError:
    print("Trying to Install required module: requests\n")
    os.system('python -m pip install openpyxl')
try:
    import shutil
except ImportError:
    print("Trying to Install required module: requests\n")
    os.system('python -m pip install shutil')
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# In[ ]:


#input assignment title, date, and filepath
title = input('\nTitle: ')
date = input('\nDate: ')
fpath = input('\nFile Path: ')


# In[ ]:


#path = (r"C:\Users\alexa\OneDrive\Documents\Indiana MSIS\MSIS FA 21\TA things\GRC\\" + title + "\\")

path = ( fpath + '\\' + title + '\\')

# Check whether the specified path exists or not
isExist = os.path.exists(path)

if not isExist:
  
  # Create a new directory because it does not exist 
    os.makedirs(path)
    print("\nThe new directory is created\n")


# In[ ]:


#find or create our workbook
# try:
    # gb = load_workbook(path + "(" + title + " - GradeSummary).xlsx")
    # gs = gb.active
# except FileNotFoundError or PermissionError:
    # gb = Workbook()
    # gs = gb.active

#find or create our workbook
gbExist = os.path.exists(path + "(" + title + " - GradeSummary).xlsx")
if gbExist == False:
    gb = Workbook()
    gs = gb.active
    #enter headers into grading summary
    gs['A1'] = 'Name'
    gs['B1'] = 'File Reference'
    gs['C1'] = 'Grade'
    gs['D1'] = 'Raw Score'
    gs['A1'].font = Font(bold=True)
    gs['B1'].font = Font(bold=True)
    gs['C1'].font = Font(bold=True)
    gs['D1'].font = Font(bold=True)
    gb.save((path + "(" + title + " - GradeSummary).xlsx"))

elif gbExist == True:
    while True:
        try:
            gb = load_workbook(path + "(" + title + " - GradeSummary).xlsx")
            break
        except PermissionError:
            print('\nGrade summary cannot be loaded because it is currently in use')
            print('Please close the file before continuing\n')
            tryAgain = input('Try again? (y/n): ')
            
            if tryAgain == 'n':
                quit()
            continue

    gs = gb.active
    

   
                


    
#gs = gb.active

# In[ ]:


#function to remove empty rows from grading summary
def remove(gs):
  # iterate the sheet by rows
  for row in gs.iter_rows():
  
    # all() return False if all of the row value is None
    if not all(cell.value for cell in row):
  
      # delete the empty row
      gs.delete_rows(row[0].row, 1)
  
      # recursively call the remove() with modified sheet data
      remove(gs)
  
      return


# In[ ]:


hotkeys = {'a': 1,
          'b': .75,
          'c':.4}


# In[ ]:


hotkeyList = ['a','b','c']


# In[ ]:


#enter headers into grading summary
gs['A1'] = 'Name'
gs['B1'] = 'File Reference'
gs['C1'] = 'Grade'
gs['D1'] = 'Raw Score'
gs['A1'].font = Font(bold=True)
gs['B1'].font = Font(bold=True)
gs['C1'].font = Font(bold=True)
gs['D1'].font = Font(bold=True)

#set j as first blank row in excel grading summary
j = int(gs.max_row) + 1
#iterate through grading summary
for i in range(j,(j+1000)):

#input student's name
#quit feature is also located here because I'm an idiot
        fname = input('First Name: ')
        if fname == 'quit':
            if __name__ == '__main__':

                for row in gs:
                    remove(gs) 
            try:
                gb.save((path + "(" + title + " - GradeSummary).xlsx"))
                raise SystemExit
            except PermissionError:
                print('Workbook could not be saved. Please close workbook and try again.')
                continue
        else:
            lname = input('Last Name: ')
        if lname == 'quit':
            if __name__ == '__main__':

                # iterate the sheet
                for row in gs:
                    remove(gs) 
            try:     
                gb.save((path + "(" + title + " - GradeSummary).xlsx"))
                raise SystemExit
            except PermissionError:
                print('Workbook could not be saved. Please close workbook and try again.')
                continue
            
#check for submission
        sub = input('Submitted assignment? (y/n): ')
        while sub not in ('y','n'):
                    print('Invalid entry. Please enter "y" or "n"')
                    sub = input('Submitted assignment? (y/n): ')
        if sub == 'y':

#create copy of rubric to fill out with student's info/grades
            shutil.copy("briefRubric.xlsx", (path + lname + " " + fname + " " + title + ".xlsx"))

            wb = openpyxl.load_workbook(filename=(path + lname + " " + fname + " " + title + ".xlsx"))
            ws = wb.active

    #enter header info        
            ws['C3'] = title
            ws['C4'] = date
            ws['C5'] = fname + " " + lname

    #e9        
            #input validation 
            e9_input = (input('Provided a single-page Executive Briefing summary (/.5): '))
            while e9_input not in (hotkeyList):
                e9_input = (input('\nInput options: a, b, or c\n\nProvided a single-page Executive Briefing summary (/.5): '))
                if e9_input in (hotkeyList):
                    break
            
            #adjust variable for rubric value
            e9 = float(hotkeys[e9_input])*.5
            print(e9)
            
            if e9 == .5:
                ws['F9'] = 'Fully met criteria'
            elif .21 < e9 < .5:
                ws['F9'] = 'Partially met criteria'
            else:
                ws['F9'] = 'Did not meet criteria'
            
            if e9 < .5:
                ws['B10'] = input('Comment: ')
            ws['E9'] = e9
    #e11        
            #input validation    
            e11_input = (input('Covered the content asked for in the Case Study Executive Briefing assignment (/2): '))
            while e11_input not in (hotkeyList):
                e11_input = (input('\nInput options: a, b, or c\n\nCovered the content asked for in the Case Study Executive Briefing assignment (/2): '))
                if e11_input in (hotkeyList):
                    break
            
            #adjust variable for rubric value
            e11 = float(hotkeys[e11_input])*2
            print(e11)
            
            if e11 == 2:
                ws['F11'] = 'Fully met criteria'
            elif .81 < e11 < 2:
                ws['F11'] = 'Partially met criteria'
            else:
                ws['F11'] = 'Did not meet criteria'
            
            if e11 < 2:
                ws['B12'] = input('Comment: ')
            ws['E11'] = e11
    #e13        
               
            #input validation    
            e13_input = (input('Related the content to the IT GRC topic being covered (/1): '))
            while e13_input not in (hotkeyList):
                e13_input = (input('\nInput options: a, b, or c\n\nRelated the content to the IT GRC topic being covered (/1): '))
                if e13_input in (hotkeyList):
                    break
            
            #adjust variable for rubric value
            e13 = float(hotkeys[e13_input])
            print(e13)
            
            if e13 == 1:
                ws['F13'] = 'Fully met criteria'
            elif .41 < e13 < 1:
                ws['F13'] = 'Partially met criteria'
            else:
                ws['F13'] = 'Did not meet criteria'
            
            if e13 < 1:
                ws['B14'] = input('Comment: ')
            ws['E13'] = e13
    #e15        
            
            #input validation    
            e15_input = (input('Has considered the audience for the Executive briefing (/1): '))
            while e15_input not in (hotkeyList):
                e15_input = (input('\nInput options: a, b, or c\n\nHas considered the audience for the Executive briefing (/1): '))
                if e15_input in (hotkeyList):
                    break
            
            #adjust variable for rubric value
            e15 = float(hotkeys[e15_input])
            print(e15)
                
            if e15 == 1:
                ws['F15'] = 'Fully met criteria'
            elif .41 < e15 < 1:
                ws['F15'] = 'Partially met criteria'
            else:
                ws['F15'] = 'Did not meet criteria'
            
            if e15 < 1:
                ws['B16'] = input('Comment: ')
            ws['E15'] = e15
    #e17        

            #input validation    
            e17_input = (input('Provided a professionally formatted one-pager (e.g.: font, colors, alignment, etc. consistent) (/1): '))
            while e17_input not in (hotkeyList):
                e17_input = (input('\nInput options: a, b, or c\n\nProvided a professionally formatted one-pager (e.g.: font, colors, alignment, etc. consistent) (/1): '))
                if e17_input in (hotkeyList):
                    break
            
            #adjust variable for rubric value
            e17 = float(hotkeys[e17_input])
            print(e17)
            
            if e17 == 1:
                ws['F17'] = 'Fully met criteria'
            elif .41 < e17 < 1:
                ws['F17'] = 'Partially met criteria'
            else:
                ws['F17'] = 'Did not meet criteria'
            if e17 < 1:
                ws['B18'] = input('Comment: ')
            ws['E17'] = e17
    #e19        
            
            #input validation    
            e19_input = (input('No spelling, grammar, etc. errors exist (/.5): '))
            while e19_input not in (hotkeyList):
                e19_input = (input('\nInput options: a, b, or c\n\nNo spelling, grammar, etc. errors exist (/.5): '))
                if e19_input in (hotkeyList):
                    break
            
            #adjust variable for rubric value
            e19 = float(hotkeys[e19_input])*.5
            print(e19)
            
            if e19 == .5:
                ws['F19'] = 'Fully met criteria'
            elif .21 < e19 < .5:
                ws['F19'] = 'Partially met criteria'
            else:
                ws['F19'] = 'Did not meet criteria'
            if e19 < .5:
                ws['B20'] = input('Comment: ')
            ws['E19'] = e19

    #        for j in (9, 11, 13, 15, 17, 19):
    #            cellref=ws.cell(row=j, column=5)
    #            cellref.value=input()

    #merge cells for formatting
            ws.merge_cells('E9:E10')
            ws.merge_cells('E11:E12')
            ws.merge_cells('E13:E14')
            ws.merge_cells('E15:E16')
            ws.merge_cells('E17:E18')
            ws.merge_cells('E19:E20')
            ws.merge_cells('F9:F10')
            ws.merge_cells('F11:F12')
            ws.merge_cells('F13:F14')
            ws.merge_cells('F15:F16')
            ws.merge_cells('F17:F18')
            ws.merge_cells('F19:F20')

    #save rubric
            wb.save((path + lname + " " + fname + " " + title + ".xlsx"))

    #enter student info into grading summary
            grade = round(((((e9 + e11 + e13 + e15 + e17 + e19)*(7/6))/7)*100), 2)
            print("\n" + fname + " " + lname + " received " + str(grade) + "%")
            save = input('Save this entry? (y/n): ')
            while save not in ('y','n'):
                    print('\nInvalid entry. Please enter "y" or "n"')
                    save = input('Save this entry? (y/n): ') 
            if save == 'y':
                cellref=gs.cell(row=i, column=1)
                cellref.value= fname + " " + lname
                cellref=gs.cell(row=i, column=2)
                cellref.value = lname + " " + fname + " " + title
                cellref=gs.cell(row=i, column=3)
                cellref.value = float(((e9 + e11 + e13 + e15 + e17 + e19)*(7/6))/7)
                cellref=gs.cell(row=i, column=4)
                cellref.value = float((e9 + e11 + e13 + e15 + e17 + e19)*(7/6))
                print('\nEntry saved\n')
            elif save == 'n':
                os.remove(path + lname + " " + fname + " " + title + ".xlsx")
                print('\nEntry deleted\n')
        
        else:               
            print("\n"+fname + " " + lname + ": no submission")
            save = input('\nSave this entry? (y/n): ') 
            while save not in ('y','n'):
                    print('Invalid entry. Please enter "y" or "n"')
                    save = input('Save this entry? (y/n): ') 
            if save == 'y':
                cellref=gs.cell(row=i, column=1)
                cellref.value= fname + " " + lname
                cellref=gs.cell(row=i, column=2)
                cellref.value = 'n/a'
                cellref=gs.cell(row=i, column=3)
                cellref.value = 'n/a'
                cellref=gs.cell(row=i, column=4)
                cellref.value = 'n/a'
                print('\nEntry saved\n')
            elif save == 'n':
                print('\nEntry deleted\n')
 
                



if fname != 'quit' and lname != 'quit':
    gb.save((path + "(" + title + " - GradeSummary).xlsx"))

